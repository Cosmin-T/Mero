# daily_update.py
# Daily script to check for new appointments and update the Excel file

import json
import os
import re
import time
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from load_cookies import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import Playwright, Route, sync_playwright
from util import *

logger = setup_logging()

# Storage for captured appointments
all_captured_appointments = []

# Excel file to update (use the most recent one)
EXCEL_FILE_PATTERN = "revenue_analysis_*.xlsx"


def extract_category(service_name):
    """Extract category from service name."""
    categories = {
        "Pensat": r"Pensat|ProBrows",
        "Epilat": r"Epilat",
        "Oxygenera": r"Ox[iy]gen[ae]r[ae]|Oxygen",  # Catches Oxigenera, Oxygenera, Oxigenare, Oxygenare, Oxygen Pro, etc.
        "Laminare": r"Laminare",
        "Microneedling": r"Microneedling|DermaPen",
        "ProBrows": r"ProBrows",
        "Pure Solution": r"Pure Solution",
        "Peeling": r"Peeling",
        "Tratament": r"Tratament",
    }

    for category, pattern in categories.items():
        if re.search(pattern, service_name, re.IGNORECASE):
            return category

    return "Other"


def parse_appointment(entry):
    """Parse appointment entry from API response."""
    try:
        if entry.get("type") != 0:  # Type 0 = appointment
            return None

        payload = entry.get("payload", {})
        start = entry.get("localStart", {})
        date_str = start.get("dateStr")

        if not date_str:
            return None

        booked_services = payload.get("bookedServices", [])
        appointments = []

        for service in booked_services:
            service_name = service.get("name", "")
            price = 0

            if "customPrice" in service:
                custom_price = service["customPrice"]
                if custom_price.get("type") == "Fixed":
                    amount_data = custom_price.get("fixed", {}).get("amount", {})
                    value = amount_data.get("value", 0)
                    scale = amount_data.get("scale", 0)
                    price = value / (10**scale) if scale > 0 else value
            elif "price" in service:
                price_data = service["price"]
                if price_data.get("type") == 1:
                    price = price_data.get("fixed", 0)

            category = extract_category(service_name)

            client_name = ""
            if "client" in payload:
                client = payload["client"]
                firstname = client.get("firstname", "")
                lastname = client.get("lastname", "")
                client_name = f"{firstname} {lastname}".strip()

            appointments.append(
                {
                    "date": date_str,
                    "service": service_name,
                    "category": category,
                    "price": price,
                    "client": client_name,
                }
            )

        return appointments

    except Exception as e:
        logger.error(f"Error parsing appointment: {e}")
        return None


def handle_route(route: Route):
    """Intercept calendar API responses."""
    request = route.request
    response = route.fetch()

    url = request.url

    # Check if this is the calendar entries API
    if "calendars-entries" in url:
        logger.info(f"📡 Captured calendar API call")

        try:
            body = response.body()
            json_data = json.loads(body)

            # Parse appointments
            if "calendars" in json_data:
                for calendar in json_data["calendars"]:
                    entries = calendar.get("entries", [])
                    for entry in entries:
                        parsed_appts = parse_appointment(entry)
                        if parsed_appts:
                            all_captured_appointments.extend(parsed_appts)
                            logger.info(f"  → Parsed {len(parsed_appts)} appointments")

        except Exception as e:
            logger.error(f"Error parsing API response: {e}")

    route.fulfill(response=response)


def find_latest_excel_file():
    """Find the most recent Excel file."""
    files = list(Path(".").glob(EXCEL_FILE_PATTERN))
    if not files:
        return None

    # Sort by modification time, newest first
    files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return str(files[0])


def load_existing_appointments(excel_file):
    """Load existing appointments from Excel file."""
    if not excel_file or not os.path.exists(excel_file):
        return set()

    logger.info(f"Loading existing data from: {excel_file}")

    try:
        wb = load_workbook(excel_file)
        ws = wb["Raw Data"]

        existing = set()
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If date exists
                # Create unique key: date + client + service
                key = f"{row[0]}|{row[1]}|{row[2]}"
                existing.add(key)

        wb.close()
        logger.info(f"  Loaded {len(existing)} existing appointments")
        return existing

    except Exception as e:
        logger.error(f"Error loading existing data: {e}")
        return set()


def update_excel_file(excel_file, new_appointments):
    """Update Excel file with new appointments."""

    if not new_appointments:
        logger.info("No new appointments to add")
        return

    logger.info(f"\n=== UPDATING EXCEL FILE: {excel_file} ===\n")

    # Load existing workbook
    wb = load_workbook(excel_file)

    # Update Raw Data sheet
    ws_raw = wb["Raw Data"]

    for appt in sorted(new_appointments, key=lambda x: x["date"]):
        ws_raw.append(
            [
                appt["date"],
                appt["client"],
                appt["service"],
                appt["category"],
                appt["price"],
            ]
        )

    logger.info(f"  ✓ Added {len(new_appointments)} new appointments to Raw Data")

    # Rebuild Monthly Summary and Category Totals sheets
    rebuild_summary_sheets(wb)

    # Save
    wb.save(excel_file)
    wb.close()

    logger.info(f"\n✓✓✓ Excel file updated successfully!\n")


def rebuild_summary_sheets(wb):
    """Rebuild the Monthly Summary and Category Totals sheets from Raw Data."""

    ws_raw = wb["Raw Data"]

    # Read all appointments from Raw Data
    all_appts = []
    for row in ws_raw.iter_rows(min_row=2, values_only=True):
        if row[0]:  # If date exists
            all_appts.append(
                {
                    "date": str(row[0]),
                    "client": row[1],
                    "service": row[2],
                    "category": row[3],
                    "price": float(row[4]) if row[4] else 0,
                }
            )

    # Calculate statistics
    monthly_stats = defaultdict(lambda: defaultdict(lambda: {"count": 0, "revenue": 0}))
    category_totals = defaultdict(lambda: {"count": 0, "revenue": 0})

    for appt in all_appts:
        date_obj = datetime.strptime(appt["date"], "%Y-%m-%d")
        month_key = date_obj.strftime("%Y-%m")

        category = appt["category"]
        price = appt["price"]

        monthly_stats[month_key][category]["count"] += 1
        monthly_stats[month_key][category]["revenue"] += price

        category_totals[category]["count"] += 1
        category_totals[category]["revenue"] += price

    # Delete and recreate Monthly Summary sheet
    if "Monthly Summary" in wb.sheetnames:
        del wb["Monthly Summary"]

    ws_monthly = wb.create_sheet("Monthly Summary", 1)

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    ws_monthly.append(
        ["Month", "Category", "Appointments", "Revenue (RON)", "Avg Price (RON)"]
    )
    for col_num in range(1, 6):
        cell = ws_monthly.cell(1, col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for month in sorted(monthly_stats.keys()):
        month_obj = datetime.strptime(month, "%Y-%m")
        month_name = month_obj.strftime("%B %Y")

        categories = monthly_stats[month]

        for category, stats in sorted(
            categories.items(), key=lambda x: x[1]["revenue"], reverse=True
        ):
            avg = stats["revenue"] / stats["count"] if stats["count"] > 0 else 0
            ws_monthly.append(
                [month_name, category, stats["count"], stats["revenue"], avg]
            )

        # Add month total
        month_total = sum(cat["revenue"] for cat in categories.values())
        month_total_count = sum(cat["count"] for cat in categories.values())

        row = ws_monthly.max_row + 1
        ws_monthly.cell(row, 1, f"{month_name} TOTAL")
        ws_monthly.cell(row, 2, "")
        ws_monthly.cell(row, 3, month_total_count)
        ws_monthly.cell(row, 4, month_total)
        ws_monthly.cell(row, 5, "")

        for col in range(1, 6):
            ws_monthly.cell(row, col).font = Font(bold=True)
            ws_monthly.cell(row, col).fill = PatternFill(
                start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
            )

        ws_monthly.append(["", "", "", "", ""])

    # Auto-adjust columns
    for col_num in range(1, 6):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws_monthly[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_monthly.column_dimensions[column_letter].width = min(max_length + 2, 30)

    logger.info(f"  ✓ Rebuilt Monthly Summary sheet")

    # Delete and recreate Category Totals sheet
    if "Category Totals" in wb.sheetnames:
        del wb["Category Totals"]

    ws_totals = wb.create_sheet("Category Totals", 2)

    ws_totals.append(
        [
            "Category",
            "Total Appointments",
            "Total Revenue (RON)",
            "Avg Price (RON)",
            "% of Revenue",
        ]
    )
    for col_num in range(1, 6):
        cell = ws_totals.cell(1, col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    grand_total_revenue = sum(cat["revenue"] for cat in category_totals.values())

    for category, stats in sorted(
        category_totals.items(), key=lambda x: x[1]["revenue"], reverse=True
    ):
        avg = stats["revenue"] / stats["count"] if stats["count"] > 0 else 0
        percent = (
            (stats["revenue"] / grand_total_revenue * 100)
            if grand_total_revenue > 0
            else 0
        )
        ws_totals.append(
            [category, stats["count"], stats["revenue"], avg, f"{percent:.1f}%"]
        )

    # Add grand total
    grand_total_count = sum(cat["count"] for cat in category_totals.values())
    row = ws_totals.max_row + 1
    ws_totals.cell(row, 1, "GRAND TOTAL")
    ws_totals.cell(row, 2, grand_total_count)
    ws_totals.cell(row, 3, grand_total_revenue)
    ws_totals.cell(row, 4, "")
    ws_totals.cell(row, 5, "100.0%")

    for col in range(1, 6):
        ws_totals.cell(row, col).font = Font(bold=True, color="FFFFFF", size=12)
        ws_totals.cell(row, col).fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )

    # Auto-adjust columns
    for col_num in range(1, 6):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws_totals[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_totals.column_dimensions[column_letter].width = min(max_length + 2, 30)

    logger.info(f"  ✓ Rebuilt Category Totals sheet")


def run(playwright: Playwright) -> None:
    logger.info("\n\n" + "=" * 70)
    logger.info("DAILY APPOINTMENT UPDATE")
    logger.info(f"Run time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 70)

    # Find existing Excel file
    excel_file = find_latest_excel_file()

    if not excel_file:
        logger.error(
            "No existing Excel file found! Run historical_revenue_analysis_v2.py first."
        )
        return

    logger.info(f"Using Excel file: {excel_file}")

    # Load existing appointments
    existing_keys = load_existing_appointments(excel_file)

    # Launch browser
    context = playwright.chromium.launch_persistent_context(
        "./chrome-data",
        headless=False,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--disable-dev-shm-usage",
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--ignore-certificate-errors",
        ],
        viewport={"width": 3024, "height": 1440},
        java_script_enabled=True,
        bypass_csp=True,
        ignore_https_errors=True,
    )

    context.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
    """)

    load_cookies_json(context)
    context.set_default_timeout(15000)

    page = context.new_page()

    # Enable request interception
    logger.info("Enabling API interception...")
    page.route("**/*", handle_route)

    page.goto(URL)
    page.wait_for_load_state("networkidle")

    # Login check
    login_identifier = (
        page.locator("div")
        .filter(has_text=re.compile(r"^Utilizează numărul de mobil$"))
        .nth(1)
    )

    if login_identifier.is_visible():
        logger.info("Login required - please run this manually first to login")
        page.close()
        context.close()
        return
    else:
        logger.info("Already logged in")

    page.wait_for_load_state("networkidle")
    time.sleep(3)  # Wait for calendar to load

    logger.info(
        f"\n✓ Captured {len(all_captured_appointments)} appointments from current week\n"
    )

    # Filter out existing appointments
    new_appointments = []

    for appt in all_captured_appointments:
        key = f"{appt['date']}|{appt['client']}|{appt['service']}"
        if key not in existing_keys:
            new_appointments.append(appt)

    logger.info(f"Found {len(new_appointments)} NEW appointments")

    if new_appointments:
        for appt in new_appointments:
            logger.info(
                f"  + {appt['date']} - {appt['client']} - {appt['service']} - {appt['price']} RON"
            )

    # Update Excel file
    update_excel_file(excel_file, new_appointments)

    page.close()
    context.close()


if __name__ == "__main__":
    with sync_playwright() as playwright:
        run(playwright)
