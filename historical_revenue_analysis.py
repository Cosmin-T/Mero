# historical_revenue_analysis_v2.py
# Extract historical data by navigating the calendar UI and intercepting API responses
# Then automatically enhance with comprehensive summary sheets

import json
import os
import re
import subprocess
import sys
import time
from collections import defaultdict
from datetime import datetime, timedelta

from load_cookies import *
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import Playwright, Route, sync_playwright
from util import *

logger = setup_logging()

# Storage for captured appointments
all_captured_appointments = []


def extract_category(service_name):
    """Extract category from service name."""
    categories = {
        "Pensat": r"Pensat",
        "ProBrows": r"ProBrows",
        "Epilat": r"Epilat",
        "Oxygenera": r"Ox[iy]gen[ae]r[ae]|Oxygen",  # Catches Oxigenera, Oxygenera, Oxigenare, Oxygenare, Oxygen Pro, etc.
        "Laminare": r"Laminare",
        "Microneedling": r"Microneedling",
        "DermaPen": r"DermaPen",
        "Peeling": r"Peeling",
        "Tratament": r"Tratament",
        "Pure Solution": r"Pure Solution",
    }

    for category, pattern in categories.items():
        if re.search(pattern, service_name, re.IGNORECASE):
            return category

    return "Other"


def get_mega_category(individual_category):
    """Get mega-category for an individual category."""
    mega_categories = {
        "Sprâncene": ["Pensat", "ProBrows", "Laminare"],
        "Epilare": ["Epilat"],
        "Tratamente": [
            "Oxygenera",
            "Microneedling",
            "DermaPen",
            "Peeling",
            "Tratament",
            "Pure Solution",
        ],
    }

    for mega_cat, individual_cats in mega_categories.items():
        if individual_category in individual_cats:
            return mega_cat

    return "Altele"


def parse_appointment(entry):
    """Parse appointment entry from API response."""
    try:
        if entry.get("type") != 0:  # Type 0 = appointment
            return None

        payload = entry.get("payload", {})
        start = entry.get("localStart", {})
        date_str = start.get("dateStr")

        if not date_str:
            return None

        booked_services = payload.get("bookedServices", [])
        appointments = []

        for service in booked_services:
            service_name = service.get("name", "")
            price = 0

            if "customPrice" in service:
                custom_price = service["customPrice"]
                if custom_price.get("type") == "Fixed":
                    amount_data = custom_price.get("fixed", {}).get("amount", {})
                    value = amount_data.get("value", 0)
                    scale = amount_data.get("scale", 0)
                    price = value / (10**scale) if scale > 0 else value
            elif "price" in service:
                price_data = service["price"]
                if price_data.get("type") == 1:
                    price = price_data.get("fixed", 0)

            category = extract_category(service_name)

            client_name = ""
            if "client" in payload:
                client = payload["client"]
                firstname = client.get("firstname", "")
                lastname = client.get("lastname", "")
                client_name = f"{firstname} {lastname}".strip()

            appointments.append(
                {
                    "date": date_str,
                    "service": service_name,
                    "category": category,
                    "mega_category": get_mega_category(category),
                    "price": price,
                    "client": client_name,
                }
            )

        return appointments

    except Exception as e:
        logger.error(f"Error parsing appointment: {e}")
        return None


def handle_route(route: Route):
    """Intercept calendar API responses."""
    request = route.request
    response = route.fetch()

    url = request.url

    # Check if this is the calendar entries API
    if "calendars-entries" in url:
        logger.info(f"📡 Captured calendar API call")

        try:
            body = response.body()
            json_data = json.loads(body)

            # Parse appointments
            if "calendars" in json_data:
                for calendar in json_data["calendars"]:
                    entries = calendar.get("entries", [])
                    for entry in entries:
                        parsed_appts = parse_appointment(entry)
                        if parsed_appts:
                            all_captured_appointments.extend(parsed_appts)
                            logger.info(f"  → Parsed {len(parsed_appts)} appointments")

        except Exception as e:
            logger.error(f"Error parsing API response: {e}")

    route.fulfill(response=response)


def run(playwright: Playwright) -> None:
    logger.info("\n\n" + "=" * 70)
    logger.info("HISTORICAL REVENUE ANALYSIS V2")
    logger.info("=" * 70)

    context = playwright.chromium.launch_persistent_context(
        "./chrome-data",
        headless=False,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--disable-dev-shm-usage",
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--ignore-certificate-errors",
        ],
        viewport={"width": 3024, "height": 1440},
        java_script_enabled=True,
        bypass_csp=True,
        ignore_https_errors=True,
    )

    context.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
    """)

    load_cookies_json(context)
    context.set_default_timeout(15000)

    page = context.new_page()

    # Enable request interception
    logger.info("Enabling API interception...")
    page.route("**/*", handle_route)

    page.goto(URL)
    page.wait_for_load_state("networkidle")

    # Login if needed (condensed version)
    login_identifier = (
        page.locator("div")
        .filter(has_text=re.compile(r"^Utilizează numărul de mobil$"))
        .nth(1)
    )

    if login_identifier.is_visible():
        logger.info("Logging in...")
        # [Login code omitted for brevity - same as before]
    else:
        logger.info("Already logged in")

    page.wait_for_load_state("networkidle")
    time.sleep(2)

    logger.info("\n=== NAVIGATING TO PREVIOUS WEEKS ===\n")

    # TEST: Navigate back 2 weeks only - change to 15 for full Sept-Dec data
    weeks_to_go_back = 53

    for week_num in range(weeks_to_go_back):
        logger.info(
            f"Week {week_num + 1}/{weeks_to_go_back}: Clicking previous week..."
        )

        try:
            page.locator(".css-175oi2r.r-1awozwy.r-13awgt0 > div > div").first.click()
            page.wait_for_load_state("networkidle", timeout=5000)
            time.sleep(2)  # Wait for API call to complete
            logger.info(
                f"  Total appointments captured so far: {len(all_captured_appointments)}"
            )
        except Exception as e:
            logger.error(f"Error navigating to week {week_num + 1}: {e}")
            break

    logger.info(
        f"\n=== TOTAL APPOINTMENTS CAPTURED: {len(all_captured_appointments)} ===\n"
    )

    # Calculate statistics
    monthly_stats = defaultdict(lambda: defaultdict(lambda: {"count": 0, "revenue": 0}))
    monthly_mega_stats = defaultdict(
        lambda: defaultdict(lambda: {"count": 0, "revenue": 0})
    )
    category_totals = defaultdict(lambda: {"count": 0, "revenue": 0})
    mega_category_totals = defaultdict(lambda: {"count": 0, "revenue": 0})

    for appt in all_captured_appointments:
        date_obj = datetime.strptime(appt["date"], "%Y-%m-%d")
        month_key = date_obj.strftime("%Y-%m")

        category = appt["category"]
        mega_category = appt["mega_category"]
        price = appt["price"]

        monthly_stats[month_key][category]["count"] += 1
        monthly_stats[month_key][category]["revenue"] += price

        monthly_mega_stats[month_key][mega_category]["count"] += 1
        monthly_mega_stats[month_key][mega_category]["revenue"] += price

        category_totals[category]["count"] += 1
        category_totals[category]["revenue"] += price

        mega_category_totals[mega_category]["count"] += 1
        mega_category_totals[mega_category]["revenue"] += price

    # Generate report
    report_lines = [
        "\n" + "=" * 70,
        "MONTHLY REVENUE REPORT (TEST - 2 WEEKS)",
        "=" * 70,
    ]

    for month in sorted(monthly_stats.keys()):
        month_obj = datetime.strptime(month, "%Y-%m")
        report_lines.append(f"\n### {month_obj.strftime('%B %Y')} ###\n")

        categories = monthly_stats[month]
        month_total = sum(cat["revenue"] for cat in categories.values())

        for category, stats in sorted(
            categories.items(), key=lambda x: x[1]["revenue"], reverse=True
        ):
            avg = stats["revenue"] / stats["count"] if stats["count"] > 0 else 0
            report_lines.append(
                f"  {category:20} : {stats['count']:4} appts | {stats['revenue']:8.2f} RON | Avg: {avg:6.2f} RON"
            )

        report_lines.append(f"  {'-' * 68}")
        report_lines.append(f"  {'MONTH TOTAL':20} : {month_total:8.2f} RON\n")

    report_text = "\n".join(report_lines)
    print(report_text)

    with open("test_revenue_report.txt", "w", encoding="utf-8") as f:
        f.write(report_text)

    logger.info("\n✓ Test report saved to test_revenue_report.txt")

    # === EXCEL EXPORT ===
    logger.info("\n=== CREATING EXCEL FILE ===\n")

    wb = Workbook()

    # Sheet 1: Raw Data
    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    # Headers
    headers = ["Date", "Client", "Service", "Category", "Mega Category", "Price (RON)"]
    ws_raw.append(headers)

    # Style headers
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws_raw.cell(1, col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data
    for appt in sorted(all_captured_appointments, key=lambda x: x["date"]):
        ws_raw.append(
            [
                appt["date"],
                appt["client"],
                appt["service"],
                appt["category"],
                appt["mega_category"],
                appt["price"],
            ]
        )

    # Apply number formatting to price column (column 6)
    for row in range(2, ws_raw.max_row + 1):
        price_cell = ws_raw.cell(row=row, column=6)
        if price_cell.value is not None:
            price_cell.number_format = '#,##0.00" RON"'

    # Auto-adjust column widths
    for col_num in range(1, 7):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws_raw[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_raw.column_dimensions[column_letter].width = min(max_length + 2, 50)

    logger.info(f"  ✓ Raw Data sheet: {len(all_captured_appointments)} appointments")

    # Sheet 2: Monthly Summary
    ws_monthly = wb.create_sheet("Monthly Summary")

    # Headers
    ws_monthly.append(
        ["Month", "Category", "Appointments", "Revenue (RON)", "Avg Price (RON)"]
    )
    for col_num in range(1, 6):
        cell = ws_monthly.cell(1, col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add monthly data
    for month in sorted(monthly_stats.keys()):
        month_obj = datetime.strptime(month, "%Y-%m")
        month_name = month_obj.strftime("%B %Y")

        categories = monthly_stats[month]

        for category, stats in sorted(
            categories.items(), key=lambda x: x[1]["revenue"], reverse=True
        ):
            avg = stats["revenue"] / stats["count"] if stats["count"] > 0 else 0
            ws_monthly.append(
                [month_name, category, stats["count"], stats["revenue"], avg]
            )

        # Add month total
        month_total = sum(cat["revenue"] for cat in categories.values())
        month_total_count = sum(cat["count"] for cat in categories.values())

        row = ws_monthly.max_row + 1
        ws_monthly.cell(row, 1, f"{month_name} TOTAL")
        ws_monthly.cell(row, 2, "")
        ws_monthly.cell(row, 3, month_total_count)
        ws_monthly.cell(row, 4, month_total)
        ws_monthly.cell(row, 5, "")

        # Bold the total row
        for col in range(1, 6):
            ws_monthly.cell(row, col).font = Font(bold=True)
            ws_monthly.cell(row, col).fill = PatternFill(
                start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
            )

        # Add empty row
        ws_monthly.append(["", "", "", "", ""])

    # Apply number formatting to revenue and avg price columns (columns 4 and 5)
    for row in range(2, ws_monthly.max_row + 1):
        revenue_cell = ws_monthly.cell(row=row, column=4)
        avg_price_cell = ws_monthly.cell(row=row, column=5)

        if revenue_cell.value is not None:
            revenue_cell.number_format = '#,##0.00" RON"'
        if avg_price_cell.value is not None:
            avg_price_cell.number_format = '#,##0.00" RON"'

    # Auto-adjust column widths
    for col_num in range(1, 6):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws_monthly[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_monthly.column_dimensions[column_letter].width = min(max_length + 2, 30)

    logger.info(f"  ✓ Monthly Summary sheet created")

    # Sheet 3: Category Totals
    ws_totals = wb.create_sheet("Category Totals")

    # Headers
    ws_totals.append(
        [
            "Category",
            "Total Appointments",
            "Total Revenue (RON)",
            "Avg Price (RON)",
            "% of Revenue",
        ]
    )
    for col_num in range(1, 6):
        cell = ws_totals.cell(1, col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Calculate grand total
    grand_total_revenue = sum(cat["revenue"] for cat in category_totals.values())

    # Add category data
    for category, stats in sorted(
        category_totals.items(), key=lambda x: x[1]["revenue"], reverse=True
    ):
        avg = stats["revenue"] / stats["count"] if stats["count"] > 0 else 0
        percent = (
            (stats["revenue"] / grand_total_revenue * 100)
            if grand_total_revenue > 0
            else 0
        )

        ws_totals.append(
            [category, stats["count"], stats["revenue"], avg, f"{percent:.1f}%"]
        )

    # Add grand total
    grand_total_count = sum(cat["count"] for cat in category_totals.values())
    row = ws_totals.max_row + 1
    ws_totals.cell(row, 1, "GRAND TOTAL")
    ws_totals.cell(row, 2, grand_total_count)
    ws_totals.cell(row, 3, grand_total_revenue)
    ws_totals.cell(row, 4, "")
    ws_totals.cell(row, 5, "100.0%")

    # Bold the total row
    for col in range(1, 6):
        ws_totals.cell(row, col).font = Font(bold=True, size=12)
        ws_totals.cell(row, col).fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        ws_totals.cell(row, col).font = Font(bold=True, color="FFFFFF", size=12)

    # Apply number formatting to revenue and avg price columns (columns 3 and 4)
    for row in range(2, ws_totals.max_row + 1):
        revenue_cell = ws_totals.cell(row=row, column=3)
        avg_price_cell = ws_totals.cell(row=row, column=4)

        if revenue_cell.value is not None:
            revenue_cell.number_format = '#,##0.00" RON"'
        if avg_price_cell.value is not None:
            avg_price_cell.number_format = '#,##0.00" RON"'

    # Auto-adjust column widths
    for col_num in range(1, 6):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws_totals[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_totals.column_dimensions[column_letter].width = min(max_length + 2, 30)

    logger.info(f"  ✓ Category Totals sheet created")

    # Sheet 4: Mega Category Totals
    ws_mega_totals = wb.create_sheet("Mega Category Totals")

    # Headers
    ws_mega_totals.append(
        [
            "Mega Category",
            "Total Appointments",
            "Total Revenue (RON)",
            "Avg Price (RON)",
            "% of Revenue",
        ]
    )
    for col_num in range(1, 6):
        cell = ws_mega_totals.cell(1, col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Calculate grand total
    grand_total_revenue = sum(cat["revenue"] for cat in mega_category_totals.values())

    # Add mega category data
    for mega_category, stats in sorted(
        mega_category_totals.items(), key=lambda x: x[1]["revenue"], reverse=True
    ):
        avg = stats["revenue"] / stats["count"] if stats["count"] > 0 else 0
        percent = (
            (stats["revenue"] / grand_total_revenue * 100)
            if grand_total_revenue > 0
            else 0
        )

        ws_mega_totals.append(
            [mega_category, stats["count"], stats["revenue"], avg, f"{percent:.1f}%"]
        )

    # Add grand total
    grand_total_count = sum(cat["count"] for cat in mega_category_totals.values())
    row = ws_mega_totals.max_row + 1
    ws_mega_totals.cell(row, 1, "GRAND TOTAL")
    ws_mega_totals.cell(row, 2, grand_total_count)
    ws_mega_totals.cell(row, 3, grand_total_revenue)
    ws_mega_totals.cell(row, 4, "")
    ws_mega_totals.cell(row, 5, "100.0%")

    # Bold the total row
    for col in range(1, 6):
        ws_mega_totals.cell(row, col).font = Font(bold=True, size=12)
        ws_mega_totals.cell(row, col).fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        ws_mega_totals.cell(row, col).font = Font(bold=True, color="FFFFFF", size=12)

    # Apply number formatting to revenue and avg price columns (columns 3 and 4)
    for row_num in range(2, ws_mega_totals.max_row + 1):
        revenue_cell = ws_mega_totals.cell(row=row_num, column=3)
        avg_price_cell = ws_mega_totals.cell(row=row_num, column=4)

        if revenue_cell.value is not None:
            revenue_cell.number_format = '#,##0.00" RON"'
        if avg_price_cell.value is not None:
            avg_price_cell.number_format = '#,##0.00" RON"'

    # Auto-adjust column widths
    for col_num in range(1, 6):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws_mega_totals[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_mega_totals.column_dimensions[column_letter].width = min(max_length + 2, 30)

    logger.info(f"  ✓ Mega Category Totals sheet created")

    # Save Excel file with consistent name
    excel_filename = "revenue_analysis_current.xlsx"
    wb.save(excel_filename)

    logger.info(f"\n✓✓✓ Excel file saved: {excel_filename}\n")

    # Run the summary enhancer on the created file
    logger.info("=== RUNNING EXCEL SUMMARY ENHANCER ===")
    try:
        # Get the directory of this script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        summary_script = os.path.join(script_dir, "excel_summary_enhancer.py")

        # Check if summary script exists
        if os.path.exists(summary_script):
            logger.info(f"Running summary enhancer: {summary_script}")

            # Run the summary enhancer
            result = subprocess.run(
                [sys.executable, summary_script],
                cwd=script_dir,
                capture_output=True,
                text=True,
            )

            if result.returncode == 0:
                logger.info("✓ Summary enhancer completed successfully")
                logger.info(f"Output: {result.stdout}")
            else:
                logger.error(f"✗ Summary enhancer failed with code {result.returncode}")
                logger.error(f"Error: {result.stderr}")
        else:
            logger.warning(f"Summary enhancer script not found: {summary_script}")
            logger.warning(
                "Please ensure excel_summary_enhancer.py is in the same directory"
            )

    except Exception as e:
        logger.error(f"Error running summary enhancer: {e}")
        import traceback

        traceback.print_exc()

    page.close()
    context.close()


if __name__ == "__main__":
    with sync_playwright() as playwright:
        run(playwright)
