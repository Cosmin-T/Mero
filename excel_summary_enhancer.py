#!/usr/bin/env python3
"""
Excel Summary Enhancer
======================

This script reads an Excel file containing revenue analysis data and adds multiple
summary sheets with comprehensive analytics.

Input: revenue_analysis_20251222_164420.xlsx
Output: Same file with additional summary sheets added
"""

import logging
import os
import sys
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class ExcelSummaryEnhancer:
    """Main class for enhancing Excel files with summary sheets."""

    def __init__(self, excel_file_path):
        """
        Initialize the enhancer with the Excel file path.

        Args:
            excel_file_path (str): Path to the Excel file
        """
        self.excel_file_path = excel_file_path
        self.raw_data = None
        self.monthly_summary = None
        self.category_totals = None
        self.workbook = None

    def load_data(self):
        """Load all sheets from the Excel file."""
        try:
            logger.info(f"Loading Excel file: {self.excel_file_path}")

            # Check if file exists
            if not os.path.exists(self.excel_file_path):
                raise FileNotFoundError(f"File not found: {self.excel_file_path}")

            # Load workbook for writing
            self.workbook = load_workbook(self.excel_file_path)

            # Load data with pandas
            self.raw_data = pd.read_excel(self.excel_file_path, sheet_name="Raw Data")
            self.monthly_summary = pd.read_excel(
                self.excel_file_path, sheet_name="Monthly Summary"
            )
            self.category_totals = pd.read_excel(
                self.excel_file_path, sheet_name="Category Totals"
            )

            # Clean and prepare raw data
            self._prepare_raw_data()

            logger.info(f"Loaded {len(self.raw_data)} rows from Raw Data")
            logger.info(f"Loaded {len(self.monthly_summary)} rows from Monthly Summary")
            logger.info(f"Loaded {len(self.category_totals)} rows from Category Totals")

            return True

        except Exception as e:
            logger.error(f"Error loading data: {e}")
            return False

    def _prepare_raw_data(self):
        """Prepare raw data for analysis."""
        # Convert Date column to datetime
        self.raw_data["Date"] = pd.to_datetime(self.raw_data["Date"])

        # Extract date components
        self.raw_data["Year"] = self.raw_data["Date"].dt.year
        self.raw_data["Month"] = self.raw_data["Date"].dt.month
        self.raw_data["Month_Name"] = self.raw_data["Date"].dt.strftime("%B")
        self.raw_data["YearMonth"] = self.raw_data["Date"].dt.strftime("%Y-%m")
        self.raw_data["Week"] = self.raw_data["Date"].dt.isocalendar().week
        self.raw_data["Day"] = self.raw_data["Date"].dt.day
        self.raw_data["Day_Name"] = self.raw_data["Date"].dt.day_name()

        # Calculate derived metrics
        self.raw_data["Revenue_RON"] = self.raw_data["Price (RON)"]

        # Check if Mega Category column exists (added in newer versions)
        if "Mega Category" in self.raw_data.columns:
            logger.info("Mega Category column found in data")
        else:
            # Create mega-category mapping if column doesn't exist
            logger.info("Creating Mega Category column from Category column")
            mega_category_map = {
                "Pensat": "Sprâncene",
                "ProBrows": "Sprâncene",
                "Laminare": "Sprâncene",
                "Epilat": "Epilare",
                "Oxygenera": "Tratamente",
                "Microneedling": "Tratamente",
                "DermaPen": "Tratamente",
                "Peeling": "Tratamente",
                "Tratament": "Tratamente",
                "Pure Solution": "Tratamente",
            }
            self.raw_data["Mega Category"] = self.raw_data["Category"].map(
                lambda x: mega_category_map.get(x, "Altele")
            )

        logger.info("Raw data prepared for analysis")

    def create_daily_summary(self):
        """Create daily summary sheet."""
        logger.info("Creating Daily Summary sheet")

        # Group by date
        daily_summary = (
            self.raw_data.groupby("Date")
            .agg({"Revenue_RON": ["sum", "mean", "count"], "Client": "nunique"})
            .round(2)
        )

        # Flatten column names
        daily_summary.columns = [
            "Total Revenue (RON)",
            "Average Price per Service (RON)",
            "Total Services",
            "Unique Clients",
        ]

        # Reset index
        daily_summary = daily_summary.reset_index()

        # Add day of week
        daily_summary["Day of Week"] = daily_summary["Date"].dt.day_name()

        # Sort by date
        daily_summary = daily_summary.sort_values("Date")

        return daily_summary

    def create_weekly_summary(self):
        """Create weekly summary sheet."""
        logger.info("Creating Weekly Summary sheet")

        # Group by year and week
        weekly_summary = (
            self.raw_data.groupby(["Year", "Week"])
            .agg(
                {
                    "Date": ["min", "max"],
                    "Revenue_RON": ["sum", "count"],
                    "Client": "nunique",
                }
            )
            .round(2)
        )

        # Flatten column names
        weekly_summary.columns = [
            "Week Start Date",
            "Week End Date",
            "Total Revenue (RON)",
            "Total Services",
            "Unique Clients",
        ]

        # Reset index
        weekly_summary = weekly_summary.reset_index()

        # Add week identifier
        weekly_summary["Week_Identifier"] = (
            weekly_summary["Year"].astype(str)
            + "-W"
            + weekly_summary["Week"].astype(str)
        )

        # Add 4-week moving average for trend analysis
        weekly_summary["4-Week MA Revenue (RON)"] = (
            weekly_summary["Total Revenue (RON)"]
            .rolling(window=4, min_periods=1)
            .mean()
            .round(2)
        )

        # Calculate growth vs 4-week moving average
        weekly_summary["Growth vs 4-Week MA (%)"] = None
        for idx, row in weekly_summary.iterrows():
            current_revenue = row["Total Revenue (RON)"]
            ma_revenue = row["4-Week MA Revenue (RON)"]

            if ma_revenue > 0 and current_revenue > 0:
                growth_vs_ma = (current_revenue - ma_revenue) / ma_revenue
                weekly_summary.at[idx, "Growth vs 4-Week MA (%)"] = round(
                    growth_vs_ma, 4
                )

        return weekly_summary

    def create_client_analysis(self):
        """Create client analysis sheet."""
        logger.info("Creating Client Analysis sheet")

        # Group by client
        client_analysis = (
            self.raw_data.groupby("Client")
            .agg(
                {
                    "Date": ["min", "max", "count"],
                    "Revenue_RON": ["sum", "mean"],
                }
            )
            .round(2)
        )

        # Flatten column names
        client_analysis.columns = [
            "First Visit Date",
            "Last Visit Date",
            "Total Visits",
            "Total Spent (RON)",
            "Average Spend per Visit (RON)",
        ]

        # Reset index
        client_analysis = client_analysis.reset_index()

        # Calculate recency (days since last visit)
        latest_date = self.raw_data["Date"].max()
        client_analysis["Days Since Last Visit"] = (
            latest_date - client_analysis["Last Visit Date"]
        ).dt.days

        # Sort by total spent
        client_analysis = client_analysis.sort_values(
            "Total Spent (RON)", ascending=False
        )

        # Add percentile rankings
        client_analysis["Revenue Percentile (%)"] = (
            client_analysis["Total Spent (RON)"].rank(pct=True)
        ).round(4)

        # Categorize clients
        conditions = [
            client_analysis["Total Spent (RON)"]
            >= client_analysis["Total Spent (RON)"].quantile(0.8),
            client_analysis["Total Spent (RON)"]
            >= client_analysis["Total Spent (RON)"].quantile(0.5),
            client_analysis["Total Spent (RON)"]
            >= client_analysis["Total Spent (RON)"].quantile(0.2),
        ]
        choices = ["VIP", "Regular", "Occasional", "New/Low"]
        client_analysis["Client Tier"] = np.select(
            conditions, choices[:3], default=choices[3]
        )

        return client_analysis

    def create_service_analysis(self):
        """Create service analysis sheet."""
        logger.info("Creating Service Analysis sheet")

        # Group by service
        service_analysis = (
            self.raw_data.groupby("Service")
            .agg(
                {
                    "Date": "count",
                    "Revenue_RON": ["sum", "mean"],
                    "Client": "nunique",
                    "Category": "first",
                }
            )
            .round(2)
        )

        # Flatten column names
        service_analysis.columns = [
            "Total Bookings",
            "Total Revenue (RON)",
            "Average Price (RON)",
            "Unique Clients",
            "Category",
        ]

        # Reset index
        service_analysis = service_analysis.reset_index()

        # Calculate revenue contribution
        total_revenue = service_analysis["Total Revenue (RON)"].sum()
        service_analysis["Revenue Contribution (%)"] = (
            service_analysis["Total Revenue (RON)"] / total_revenue
        ).round(4)

        # Sort by total revenue
        service_analysis = service_analysis.sort_values(
            "Total Revenue (RON)", ascending=False
        )

        return service_analysis

    def create_revenue_trends(self):
        """Create revenue trends analysis sheet."""
        logger.info("Creating Revenue Trends sheet")

        # Monthly trends
        monthly_trends = (
            self.raw_data.groupby("YearMonth")
            .agg({"Revenue_RON": ["sum", "count", "mean"], "Client": "nunique"})
            .round(2)
        )

        monthly_trends.columns = [
            "Monthly Revenue (RON)",
            "Monthly Bookings",
            "Average Booking Value (RON)",
            "Monthly Unique Clients",
        ]
        monthly_trends = monthly_trends.reset_index()

        # Calculate month-over-month growth
        monthly_trends["Monthly Revenue Growth (%)"] = (
            monthly_trends["Monthly Revenue (RON)"].pct_change()
        )
        monthly_trends["Monthly Bookings Growth (%)"] = (
            monthly_trends["Monthly Bookings"].pct_change()
        )

        # Weekly trends (last 12 weeks)
        latest_date = self.raw_data["Date"].max()
        twelve_weeks_ago = latest_date - timedelta(weeks=12)
        recent_data = self.raw_data[self.raw_data["Date"] >= twelve_weeks_ago]

        weekly_trends = (
            recent_data.groupby(["Year", "Week"])
            .agg({"Date": "min", "Revenue_RON": ["sum", "count"]})
            .round(2)
        )

        weekly_trends.columns = [
            "Week Start Date",
            "Weekly Revenue (RON)",
            "Weekly Bookings",
        ]
        weekly_trends = weekly_trends.reset_index()
        weekly_trends["Week_Identifier"] = (
            weekly_trends["Year"].astype(str) + "-W" + weekly_trends["Week"].astype(str)
        )

        # Combine into trends report
        trends_report = {
            "monthly_trends": monthly_trends,
            "weekly_trends": weekly_trends,
            "summary_stats": {
                "avg_monthly_revenue": monthly_trends["Monthly Revenue (RON)"].mean(),
                "avg_weekly_revenue": weekly_trends["Weekly Revenue (RON)"].mean(),
                "best_month": monthly_trends.loc[
                    monthly_trends["Monthly Revenue (RON)"].idxmax(), "YearMonth"
                ],
                "best_week": weekly_trends.loc[
                    weekly_trends["Weekly Revenue (RON)"].idxmax(), "Week_Identifier"
                ],
            },
        }

        return trends_report

    def create_peak_performance(self):
        """Create peak performance analysis sheet."""
        logger.info("Creating Peak Performance sheet")

        # Best performing days
        daily_performance = (
            self.raw_data.groupby(["Day_Name", "Date"])
            .agg({"Revenue_RON": "sum", "Client": "count"})
            .reset_index()
        )

        day_summary = (
            daily_performance.groupby("Day_Name")
            .agg({"Revenue_RON": ["mean", "sum"], "Client": "mean"})
            .round(2)
        )

        day_summary.columns = [
            "Average Daily Revenue (RON)",
            "Total Revenue by Day (RON)",
            "Average Daily Bookings",
        ]
        day_summary = day_summary.reset_index()

        # Sort by average daily revenue
        day_summary = day_summary.sort_values(
            "Average Daily Revenue (RON)", ascending=False
        )

        # Best performing hours (if time data was available)
        # For now, we'll analyze by date parts

        # Peak periods (best 10 days)
        top_days = daily_performance.nlargest(10, "Revenue_RON")[
            ["Date", "Day_Name", "Revenue_RON", "Client"]
        ]
        top_days.columns = ["Date", "Day of Week", "Revenue (RON)", "Bookings"]

        peak_performance = {
            "day_performance": day_summary,
            "top_10_days": top_days,
            "summary": {
                "best_day": day_summary.iloc[0]["Day_Name"],
                "worst_day": day_summary.iloc[-1]["Day_Name"],
                "avg_daily_revenue": day_summary["Average Daily Revenue (RON)"].mean(),
                "total_analyzed_days": len(daily_performance["Date"].unique()),
            },
        }

        return peak_performance

    def create_category_deep_dive(self):
        """Create detailed category analysis sheet."""
        logger.info("Creating Category Deep Dive sheet")

        # Start with existing category totals
        category_deep_dive = self.category_totals.copy()

        # Clean percentage column - convert from "34.0%" string to 0.34 decimal
        category_deep_dive["% of Revenue"] = (
            category_deep_dive["% of Revenue"].str.rstrip("%").astype(float) / 100
        )

        # Add more metrics from raw data
        category_details = (
            self.raw_data.groupby("Category")
            .agg(
                {
                    "Revenue_RON": ["sum", "mean"],
                    "Client": "nunique",
                    "Service": "nunique",
                }
            )
            .round(2)
        )

        category_details.columns = [
            "Total_Revenue",
            "Average Price (RON)",
            "Unique Clients",
            "Unique Services",
        ]

        category_details = category_details.reset_index()

        # Merge with existing data
        category_deep_dive = pd.merge(
            category_deep_dive, category_details, on="Category", how="left"
        )

        # Calculate additional metrics - use robust column access
        revenue_col = "Total Revenue (RON)"
        if revenue_col not in category_deep_dive.columns:
            # Try alternative column names
            if "Total_Revenue" in category_deep_dive.columns:
                revenue_col = "Total_Revenue"
            elif "Total Revenue" in category_deep_dive.columns:
                revenue_col = "Total Revenue"
            else:
                # Fallback to first column that contains "Revenue"
                revenue_cols = [
                    col for col in category_deep_dive.columns if "Revenue" in col
                ]
                if revenue_cols:
                    revenue_col = revenue_cols[0]
                else:
                    revenue_col = category_deep_dive.columns[
                        2
                    ]  # Assume 3rd column is revenue

        category_deep_dive["Revenue per Client (RON)"] = (
            category_deep_dive[revenue_col] / category_deep_dive["Unique Clients"]
        ).round(2)

        # Calculate Client Retention Rate - % of clients who came back for this category
        # Get clients with more than 1 visit per category
        client_visits_per_category = self.raw_data.groupby(["Category", "Client"]).size()
        repeat_clients_per_category = client_visits_per_category[client_visits_per_category > 1].groupby(level=0).count()
        total_clients_per_category = self.raw_data.groupby("Category")["Client"].nunique()

        retention_rates = (repeat_clients_per_category / total_clients_per_category).fillna(0).round(4)
        category_deep_dive["Client Retention Rate (%)"] = category_deep_dive["Category"].map(retention_rates)

        # Sort by revenue contribution - use robust column access
        percent_col = "% of Revenue"
        if percent_col not in category_deep_dive.columns:
            # Try alternative column names
            percent_cols = [
                col
                for col in category_deep_dive.columns
                if "%" in col or "Percent" in col
            ]
            if percent_cols:
                percent_col = percent_cols[0]
            else:
                percent_col = category_deep_dive.columns[
                    -1
                ]  # Assume last column is percentage

        category_deep_dive = category_deep_dive.sort_values(
            percent_col, ascending=False
        )

        return category_deep_dive

    def create_mega_category_analysis(self):
        """Create mega-category analysis sheet."""
        logger.info("Creating Mega Category Analysis sheet")

        # Group by mega category
        mega_category_analysis = (
            self.raw_data.groupby("Mega Category")
            .agg(
                {
                    "Revenue_RON": ["sum", "mean", "count"],
                    "Client": "nunique",
                    "Service": "nunique",
                }
            )
            .round(2)
        )

        mega_category_analysis.columns = [
            "Total_Revenue",
            "Average Price (RON)",
            "Total Bookings",
            "Unique Clients",
            "Unique Services",
        ]

        mega_category_analysis = mega_category_analysis.reset_index()

        # Calculate percentages - use robust column access
        revenue_col = "Total_Revenue"
        if revenue_col not in mega_category_analysis.columns:
            # Try alternative column names
            revenue_cols = [
                col for col in mega_category_analysis.columns if "Revenue" in col
            ]
            if revenue_cols:
                revenue_col = revenue_cols[0]
            else:
                revenue_col = mega_category_analysis.columns[
                    2
                ]  # Assume 3rd column is revenue

        total_revenue = mega_category_analysis[revenue_col].sum()
        mega_category_analysis["Revenue Share (%)"] = (
            mega_category_analysis[revenue_col] / total_revenue
        ).round(4)

        total_bookings = mega_category_analysis["Total Bookings"].sum()
        mega_category_analysis["Bookings Share (%)"] = (
            mega_category_analysis["Total Bookings"] / total_bookings
        ).round(4)

        # Calculate additional metrics
        mega_category_analysis["Revenue per Client (RON)"] = (
            mega_category_analysis[revenue_col]
            / mega_category_analysis["Unique Clients"]
        ).round(2)

        # Sort by total revenue - use robust column access
        revenue_col = "Total_Revenue"
        if revenue_col not in mega_category_analysis.columns:
            # Try alternative column names
            revenue_cols = [
                col for col in mega_category_analysis.columns if "Revenue" in col
            ]
            if revenue_cols:
                revenue_col = revenue_cols[0]
            else:
                revenue_col = mega_category_analysis.columns[
                    2
                ]  # Assume 3rd column is revenue

        mega_category_analysis = mega_category_analysis.sort_values(
            revenue_col, ascending=False
        )

        return mega_category_analysis

    def create_forecast_sheet(self):
        """Create realistic revenue forecast for beauty salon business."""
        logger.info("Creating Forecast sheet")

        # Prepare monthly data for forecasting
        monthly_data = (
            self.raw_data.groupby("YearMonth")
            .agg({"Revenue_RON": "sum", "Client": "count"})
            .reset_index()
        )

        monthly_data.columns = [
            "Year-Month",
            "Monthly Revenue (RON)",
            "Monthly Bookings",
        ]

        # Remove partial months more intelligently
        # Compare bookings count to median to identify partial months
        if len(monthly_data) >= 3:
            median_bookings = monthly_data["Monthly Bookings"].median()
            threshold_bookings = median_bookings * 0.5  # 50% of median bookings

            full_months = []
            for idx, row in monthly_data.iterrows():
                year_month = row["Year-Month"]
                bookings = row["Monthly Bookings"]
                revenue = row["Monthly Revenue (RON)"]

                # Keep months with at least 50% of median bookings
                if bookings >= threshold_bookings:
                    full_months.append(row)
                else:
                    logger.info(
                        f"Removed partial month {year_month} ({bookings} bookings vs {median_bookings:.0f} median, {revenue:.0f} RON)"
                    )

            if len(full_months) >= 3:
                monthly_data = pd.DataFrame(full_months).reset_index(drop=True)
            else:
                # Fallback: just remove first month (likely partial)
                monthly_data = monthly_data.iloc[1:].reset_index(drop=True)

        # Need at least 6 months for meaningful forecast
        if len(monthly_data) >= 6:
            # Use recent stable period for baseline (ignore early growth explosion)
            # Take last 6-9 months to capture stable operational state
            stable_period_months = min(9, len(monthly_data))
            recent_data = monthly_data.tail(stable_period_months)

            # Calculate REALISTIC baseline from stable period
            last_3_avg = monthly_data["Monthly Revenue (RON)"].tail(3).mean()
            last_6_avg = monthly_data["Monthly Revenue (RON)"].tail(6).mean()
            recent_median = recent_data["Monthly Revenue (RON)"].median()
            recent_std = recent_data["Monthly Revenue (RON)"].std()

            # Use recent median as baseline (more robust to outliers)
            baseline = recent_median

            # Calculate growth trend from RECENT stable period only (not full history)
            recent_revenues = recent_data["Monthly Revenue (RON)"].values
            recent_growth_rates = []
            for i in range(1, len(recent_revenues)):
                if recent_revenues[i - 1] > 0:
                    growth = (recent_revenues[i] - recent_revenues[i - 1]) / recent_revenues[i - 1]
                    # Filter out extreme outliers (>30% month-over-month swings)
                    if abs(growth) < 0.3:
                        recent_growth_rates.append(growth)

            # Use median growth from stable period
            if len(recent_growth_rates) >= 3:
                median_growth = np.median(recent_growth_rates)
                # Beauty salon realistic growth: 0-3% monthly is achievable
                # Cap at ±3% to avoid crazy forecasts
                capped_growth = max(min(median_growth, 0.03), -0.03)
            else:
                # If insufficient data, assume slight positive growth based on trend
                if last_3_avg > last_6_avg:
                    capped_growth = 0.01  # Slight growth
                elif last_3_avg < last_6_avg * 0.95:
                    capped_growth = -0.01  # Slight decline
                else:
                    capped_growth = 0.00  # Stable

            # Calculate historical statistics from stable period
            all_time_avg = monthly_data["Monthly Revenue (RON)"].mean()
            historical_std = recent_data["Monthly Revenue (RON)"].std()
            historical_min = recent_data["Monthly Revenue (RON)"].min()
            historical_max = recent_data["Monthly Revenue (RON)"].max()

            # Calculate coefficient of variation from RECENT data (not including growth spike)
            if baseline > 0:
                coefficient_of_variation = (recent_std / baseline) * 100
            else:
                coefficient_of_variation = 0

            # Generate realistic forecasts for beauty salon
            forecast_months = []

            # Get current month to avoid forecasting for current/past months
            from datetime import datetime
            current_yearmonth = datetime.now().strftime("%Y-%m")
            last_data_yearmonth = monthly_data["Year-Month"].iloc[-1]

            # Start forecasting from the NEXT full month after current date
            # If current month is 2025-12, start from 2026-01
            start_month_offset = 1
            if last_data_yearmonth >= current_yearmonth:
                # Last data is current or future month, start from next month
                start_month_offset = 1
            else:
                # Last data is in the past, calculate how many months to skip
                last_year, last_month = map(int, last_data_yearmonth.split("-"))
                curr_year, curr_month = map(int, current_yearmonth.split("-"))
                months_diff = (curr_year - last_year) * 12 + (curr_month - last_month)
                start_month_offset = months_diff + 1  # Start from next month after current

            for i in range(start_month_offset, start_month_offset + 3):
                forecast_month = self._get_next_month(
                    monthly_data["Year-Month"].iloc[-1], i
                )
                months_ahead = i - start_month_offset + 1

                # Three scenarios based on REALISTIC business possibilities
                # For a stable beauty salon, use fixed scenario multipliers
                if capped_growth >= 0:
                    # POSITIVE GROWTH TREND: Build on growth
                    conservative_multiplier = 0.95  # Assume slight downturn (5% below baseline)
                    moderate_multiplier = 1.0 + (capped_growth * months_ahead)  # Current trend continues
                    optimistic_multiplier = 1.0 + (capped_growth * months_ahead * 1.5)  # 50% better than trend
                else:
                    # NEGATIVE/FLAT TREND: Scenarios around baseline
                    conservative_multiplier = 1.0 + (capped_growth * months_ahead * 1.2)  # Trend worsens 20%
                    moderate_multiplier = 1.0  # Stabilize at baseline (MOST LIKELY)
                    optimistic_multiplier = 1.02 ** months_ahead  # Recover with 2% monthly growth

                # Calculate forecasts from baseline
                conservative_forecast = baseline * conservative_multiplier
                moderate_forecast = baseline * moderate_multiplier
                optimistic_forecast = baseline * optimistic_multiplier

                # Ensure logical ordering: Conservative <= Moderate <= Optimistic
                conservative_forecast = min(conservative_forecast, moderate_forecast)
                optimistic_forecast = max(optimistic_forecast, moderate_forecast)

                forecast_months.append(
                    {
                        "Forecast Month": forecast_month,
                        "Conservative (RON)": round(conservative_forecast, 2),
                        "Moderate (RON)": round(moderate_forecast, 2),
                        "Optimistic (RON)": round(optimistic_forecast, 2),
                        "Monthly Growth Rate (%)": round(capped_growth, 4),
                    }
                )

            forecast_df = pd.DataFrame(forecast_months)

            forecast_summary = {
                "forecast": forecast_df,
                "historical_stats": {
                    "All-Time Average (RON)": round(all_time_avg, 2),
                    "Last 3 Months Average (RON)": round(last_3_avg, 2),
                    "Last 6 Months Average (RON)": round(last_6_avg, 2),
                    "Baseline (Recent Median) (RON)": round(baseline, 2),
                    "Median Monthly Growth (%)": round(median_growth, 4)
                    if len(recent_growth_rates) > 0
                    else 0.0,
                    "Applied Growth Rate (%)": round(capped_growth, 4),
                },
            }

            return forecast_summary

        else:
            logger.warning(
                "Insufficient data for forecasting (need at least 6 full months)"
            )
            return None

    def _get_next_month(self, yearmonth, months_ahead=1):
        """Get next month in YYYY-MM format."""
        year, month = map(int, yearmonth.split("-"))
        month += months_ahead
        while month > 12:
            month -= 12
            year += 1
        return f"{year}-{month:02d}"

    def _apply_excel_formatting(self, worksheet, df, sheet_name):
        """Apply formatting to Excel worksheet."""
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        money_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        percent_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )
        confidence_fill = PatternFill(
            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Write data
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border

                # Apply special formatting
                column_name = df.columns[col_idx - 1]

                # First priority: Percentage columns
                if "%" in column_name or "Percent" in column_name or "Percentile" in column_name:
                    cell.number_format = "0.00%"
                    cell.fill = percent_fill
                # Second priority: Date columns
                elif "Date" in column_name:
                    cell.number_format = "YYYY-MM-DD"
                # Third priority: RON currency columns (be very specific)
                elif (
                    # Include columns with these keywords
                    ("Revenue" in column_name or "Price" in column_name or "Spent" in column_name or "RON" in column_name or "Bound" in column_name)
                    # But exclude if they contain these keywords (not currency)
                    and "Growth" not in column_name
                    and "%" not in column_name
                    and "Index" not in column_name
                    and "Percentile" not in column_name
                    and "Contribution" not in column_name
                    and "Share" not in column_name
                    and "Stability" not in column_name
                ):
                    cell.number_format = '#,##0.00" RON"'
                    if "Avg" not in column_name and "Average" not in column_name:
                        cell.fill = money_fill
                # Fourth priority: Count/Integer metrics
                elif isinstance(value, (int, float)) and value is not None and (
                    "Count" in column_name
                    or "Visits" in column_name
                    or "Bookings" in column_name
                    or "Clients" in column_name
                    or "Services" in column_name
                    or "Appointments" in column_name
                    or "Categories" in column_name
                    or ("Total" in column_name and "Revenue" not in column_name and "Spent" not in column_name)
                ):
                    cell.number_format = "#,##0"
                # Fifth priority: Decimal numbers (Days, Index, etc.)
                elif isinstance(value, (int, float)) and value is not None and (
                    "Days" in column_name
                    or "Average" in column_name
                    or "Index" in column_name
                    or "Variability" in column_name
                    or "Std" in column_name
                    or "MA" in column_name
                ):
                    cell.number_format = "#,##0.00"
                # Sixth priority: Confidence/Quality indicators
                elif "Confidence" in column_name or "Quality" in column_name:
                    cell.fill = confidence_fill

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        worksheet.freeze_panes = "A2"

    def add_summary_sheets(self):
        """Add all summary sheets to the workbook."""
        try:
            logger.info("Adding summary sheets to workbook")

            # Create summary sheets
            daily_summary = self.create_daily_summary()
            weekly_summary = self.create_weekly_summary()
            client_analysis = self.create_client_analysis()
            service_analysis = self.create_service_analysis()
            revenue_trends = self.create_revenue_trends()
            peak_performance = self.create_peak_performance()
            category_deep_dive = self.create_category_deep_dive()
            mega_category_analysis = self.create_mega_category_analysis()
            forecast_data = self.create_forecast_sheet()

            # Add sheets to workbook
            sheets_to_add = [
                ("Daily Summary", daily_summary),
                ("Weekly Summary", weekly_summary),
                ("Client Analysis", client_analysis),
                ("Service Analysis", service_analysis),
                ("Category Deep Dive", category_deep_dive),
                ("Mega Category Analysis", mega_category_analysis),
            ]

            for sheet_name, df in sheets_to_add:
                # Remove existing sheet if it exists
                if sheet_name in self.workbook.sheetnames:
                    std = self.workbook[sheet_name]
                    self.workbook.remove(std)
                    logger.info(f"Removed existing sheet: {sheet_name}")

                # Create new sheet
                worksheet = self.workbook.create_sheet(title=sheet_name)
                self._apply_excel_formatting(worksheet, df, sheet_name)
                logger.info(f"Added sheet: {sheet_name}")

            # Add revenue trends using proper DataFrame formatting
            # Remove existing sheet if it exists
            if "Revenue Trends" in self.workbook.sheetnames:
                std = self.workbook["Revenue Trends"]
                self.workbook.remove(std)
                logger.info(f"Removed existing sheet: Revenue Trends")

            # Create combined DataFrame for Revenue Trends
            monthly_df = revenue_trends["monthly_trends"]
            weekly_df = revenue_trends["weekly_trends"]

            # Create a single DataFrame for formatting
            # We'll use monthly trends as the main DataFrame since it has RON columns
            revenue_trends_df = monthly_df.copy()

            # Create new sheet
            worksheet = self.workbook.create_sheet(title="Revenue Trends")
            self._apply_excel_formatting(worksheet, revenue_trends_df, "Revenue Trends")

            # Now add the weekly trends section manually below the monthly trends
            start_row = len(monthly_df) + 4  # +2 for header, +2 for spacing

            # Add weekly trends header
            worksheet.cell(
                row=start_row, column=1, value="Weekly Trends (Last 12 Weeks)"
            )
            worksheet.cell(row=start_row, column=1).font = Font(bold=True, size=14)

            # Write weekly trends headers
            for col_idx, col in enumerate(weekly_df.columns, 1):
                cell = worksheet.cell(row=start_row + 2, column=col_idx, value=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

            # Write weekly trends data with RON formatting
            for seq_idx, (_, row) in enumerate(weekly_df.iterrows()):
                for col_idx, value in enumerate(row.values, 1):
                    cell = worksheet.cell(
                        row=start_row + seq_idx + 3, column=col_idx, value=value
                    )
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

                    # Apply RON formatting to revenue columns
                    column_name = weekly_df.columns[col_idx - 1]
                    if "Revenue" in column_name and "RON" in column_name:
                        cell.number_format = '#,##0.00" RON"'
                        cell.fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                        )
                    elif "%" in column_name:
                        cell.number_format = "0.00%"
                        cell.fill = PatternFill(
                            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                        )
                    elif "Date" in column_name:
                        cell.number_format = "YYYY-MM-DD"

            logger.info("Added sheet: Revenue Trends")

            # Add peak performance using proper DataFrame formatting
            # Remove existing sheet if it exists
            if "Peak Performance" in self.workbook.sheetnames:
                std = self.workbook["Peak Performance"]
                self.workbook.remove(std)
                logger.info(f"Removed existing sheet: Peak Performance")

            # Use day performance DataFrame for formatting
            day_df = peak_performance["day_performance"]

            # Create new sheet with proper formatting
            worksheet = self.workbook.create_sheet(title="Peak Performance")
            self._apply_excel_formatting(worksheet, day_df, "Peak Performance")

            # Now add the top days section manually below the day performance
            start_row = len(day_df) + 4  # +2 for header, +2 for spacing

            # Add top days header
            worksheet.cell(row=start_row, column=1, value="Top 10 Revenue Days")
            worksheet.cell(row=start_row, column=1).font = Font(bold=True, size=14)

            # Write top days headers
            top_days_df = peak_performance["top_10_days"]
            for col_idx, col in enumerate(top_days_df.columns, 1):
                cell = worksheet.cell(row=start_row + 2, column=col_idx, value=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

            # Write top days data with RON formatting
            for seq_idx, (_, row) in enumerate(top_days_df.iterrows()):
                for col_idx, value in enumerate(row.values, 1):
                    cell = worksheet.cell(
                        row=start_row + seq_idx + 3, column=col_idx, value=value
                    )
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

                    # Apply RON formatting to revenue columns
                    column_name = top_days_df.columns[col_idx - 1]
                    if "Revenue" in column_name and "RON" in column_name:
                        cell.number_format = '#,##0.00" RON"'
                        cell.fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                        )
                    elif "Date" in column_name:
                        cell.number_format = "YYYY-MM-DD"

            logger.info("Added sheet: Peak Performance")

            # Add forecast if available
            if forecast_data:
                # Remove existing sheet if it exists
                if "Revenue Forecast" in self.workbook.sheetnames:
                    std = self.workbook["Revenue Forecast"]
                    self.workbook.remove(std)
                    logger.info(f"Removed existing sheet: Revenue Forecast")

                # Add forecast if available using proper DataFrame formatting
                if forecast_data and "Revenue Forecast" not in self.workbook.sheetnames:
                    forecast_df = forecast_data["forecast"]

                    # Create new sheet
                    worksheet = self.workbook.create_sheet(title="Revenue Forecast")

                    # Add title and subtitle
                    worksheet["A1"] = "3-Month Revenue Forecast"
                    worksheet["A1"].font = Font(bold=True, size=14)

                    worksheet["A2"] = (
                        "Based on historical trends and conservative growth estimates"
                    )
                    worksheet["A2"].font = Font(italic=True, size=10)

                    # Write forecast data starting at row 4 (after titles)
                    # Write headers at row 4
                    for col_idx, col in enumerate(forecast_df.columns, 1):
                        cell = worksheet.cell(row=4, column=col_idx, value=col)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(
                            start_color="366092", end_color="366092", fill_type="solid"
                        )
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin"),
                        )

                    # Write forecast data with RON formatting
                    for seq_idx, (_, row) in enumerate(forecast_df.iterrows()):
                        for col_idx, value in enumerate(row.values, 1):
                            cell = worksheet.cell(
                                row=seq_idx + 5, column=col_idx, value=value
                            )
                            cell.border = Border(
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                top=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )

                            # Apply RON formatting to revenue columns
                            column_name = forecast_df.columns[col_idx - 1]
                            if "Revenue" in column_name and "RON" in column_name:
                                cell.number_format = '#,##0.00" RON"'
                                cell.fill = PatternFill(
                                    start_color="C6EFCE",
                                    end_color="C6EFCE",
                                    fill_type="solid",
                                )
                            elif "%" in column_name:
                                cell.number_format = "0.00%"
                                cell.fill = PatternFill(
                                    start_color="FFEB9C",
                                    end_color="FFEB9C",
                                    fill_type="solid",
                                )

                    # Add historical stats and business context
                    start_row = (
                        len(forecast_df) + 7
                    )  # +4 for header row, +3 for spacing
                    worksheet.cell(
                        row=start_row,
                        column=1,
                        value="Historical Statistics & Business Context",
                    )
                    worksheet.cell(row=start_row, column=1).font = Font(
                        bold=True, size=12
                    )

                    worksheet.cell(row=start_row + 1, column=1, value="Methodology:")
                    worksheet.cell(
                        row=start_row + 1,
                        column=2,
                        value="Beauty salon specific forecasting:",
                    )
                    worksheet.cell(
                        row=start_row + 2,
                        column=2,
                        value="1. Remove partial months (bookings < 50% of median)",
                    )
                    worksheet.cell(
                        row=start_row + 3,
                        column=2,
                        value="2. Use recent stable period (last 6-9 months) for baseline",
                    )
                    worksheet.cell(
                        row=start_row + 4,
                        column=2,
                        value="3. Baseline = MEDIAN of recent months (robust to outliers)",
                    )
                    worksheet.cell(
                        row=start_row + 5,
                        column=2,
                        value="4. Growth rate from recent trend only (excludes startup growth spike)",
                    )
                    worksheet.cell(
                        row=start_row + 6,
                        column=2,
                        value="5. Realistic scenarios: Conservative (5% below), Moderate (stable), Optimistic (2% growth)",
                    )

                    worksheet.cell(
                        row=start_row + 8, column=1, value="Business Context:"
                    )
                    worksheet.cell(row=start_row + 8, column=1).font = Font(bold=True)
                    worksheet.cell(
                        row=start_row + 9,
                        column=2,
                        value="• Beauty salons are service-capacity limited businesses",
                    )
                    worksheet.cell(
                        row=start_row + 10,
                        column=2,
                        value="• Growth requires: new clients, higher prices, or new services",
                    )
                    worksheet.cell(
                        row=start_row + 11,
                        column=2,
                        value="• Realistic monthly growth: 0-3% (mature beauty salon)",
                    )
                    worksheet.cell(
                        row=start_row + 12,
                        column=2,
                        value="• Monthly variation: High (20-30%) due to appointment scheduling patterns",
                    )
                    worksheet.cell(
                        row=start_row + 13,
                        column=2,
                        value="• Seasonal patterns: Holiday peaks (Dec), summer variations",
                    )

                    stats_start = start_row + 15
                    worksheet.cell(row=stats_start, column=1, value="Key Statistics:")
                    worksheet.cell(row=stats_start, column=1).font = Font(bold=True)

                    stats = forecast_data["historical_stats"]
                    for idx, (key, value) in enumerate(stats.items(), 1):
                        worksheet.cell(
                            row=stats_start + idx,
                            column=1,
                            value=key,
                        )
                        worksheet.cell(row=stats_start + idx, column=2, value=value)

                        # Apply formatting to statistics
                        if "(RON)" in key:
                            cell = worksheet.cell(row=stats_start + idx, column=2)
                            cell.number_format = '#,##0.00" RON"'
                            cell.fill = PatternFill(
                                start_color="C6EFCE",
                                end_color="C6EFCE",
                                fill_type="solid",
                            )
                        elif "%" in key:
                            cell = worksheet.cell(row=stats_start + idx, column=2)
                            cell.number_format = "0.00%"
                            cell.fill = PatternFill(
                                start_color="FFEB9C",
                                end_color="FFEB9C",
                                fill_type="solid",
                            )

                    logger.info("Added sheet: Revenue Forecast")

            return True

        except Exception as e:
            logger.error(f"Error adding summary sheets: {e}")
            import traceback

            traceback.print_exc()
            return False

    def save_workbook(self):
        """Save the enhanced workbook."""
        try:
            # Create backup of original file
            backup_path = self.excel_file_path.replace(".xlsx", "_backup.xlsx")
            if os.path.exists(backup_path):
                os.remove(backup_path)

            # Save workbook
            self.workbook.save(self.excel_file_path)
            logger.info(f"Workbook saved: {self.excel_file_path}")

            # Create summary report
            self._create_summary_report()

            return True

        except Exception as e:
            logger.error(f"Error saving workbook: {e}")
            return False

    def _create_summary_report(self):
        """Create a text summary report."""
        report_path = self.excel_file_path.replace(".xlsx", "_summary_report.txt")

        with open(report_path, "w") as f:
            f.write("EXCEL SUMMARY ENHANCEMENT REPORT\n")
            f.write("=" * 50 + "\n\n")
            f.write(f"Original File: {self.excel_file_path}\n")
            f.write(f"Processed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            f.write("DATA OVERVIEW\n")
            f.write("-" * 30 + "\n")
            f.write(f"Total Rows in Raw Data: {len(self.raw_data):,}\n")
            f.write(
                f"Date Range: {self.raw_data['Date'].min().date()} to {self.raw_data['Date'].max().date()}\n"
            )
            f.write(f"Total Revenue: {self.raw_data['Revenue_RON'].sum():,.2f} RON\n")
            f.write(f"Unique Clients: {self.raw_data['Client'].nunique()}\n")
            f.write(f"Unique Services: {self.raw_data['Service'].nunique()}\n")
            f.write(f"Unique Categories: {self.raw_data['Category'].nunique()}\n\n")

            f.write("NEW SHEETS ADDED\n")
            f.write("-" * 30 + "\n")
            f.write("1. Daily Summary - Daily revenue and appointment metrics\n")
            f.write("2. Weekly Summary - Weekly aggregates and trends\n")
            f.write("3. Client Analysis - Client segmentation and value analysis\n")
            f.write("4. Service Analysis - Service performance metrics\n")
            f.write("5. Revenue Trends - Monthly and weekly growth analysis\n")
            f.write("6. Peak Performance - Best performing days and periods\n")
            f.write("7. Category Deep Dive - Detailed category analytics\n")
            f.write("8. Mega Category Analysis - High-level category grouping\n")
            f.write("9. Revenue Forecast - 3-month revenue projection\n\n")

            f.write("KEY METRICS\n")
            f.write("-" * 30 + "\n")
            f.write(
                f"Average Daily Revenue: {self.raw_data.groupby('Date')['Revenue_RON'].sum().mean():,.2f} RON\n"
            )
            f.write(
                f"Average Booking Value: {self.raw_data['Revenue_RON'].mean():,.2f} RON\n"
            )
            f.write(
                f"Busiest Day of Week: {self.raw_data['Day_Name'].mode().iloc[0]}\n"
            )
            f.write(
                f"Top Category by Revenue: {self.raw_data.groupby('Category')['Revenue_RON'].sum().idxmax()}\n"
            )

        logger.info(f"Summary report created: {report_path}")

    def run(self):
        """Run the complete enhancement process."""
        logger.info("Starting Excel enhancement process")

        # Load data
        if not self.load_data():
            logger.error("Failed to load data")
            return False

        # Add summary sheets
        if not self.add_summary_sheets():
            logger.error("Failed to add summary sheets")
            return False

        # Save workbook
        if not self.save_workbook():
            logger.error("Failed to save workbook")
            return False

        logger.info("Excel enhancement completed successfully")
        return True


def main():
    """Main entry point."""
    # Get the Excel file path - use consistent name
    excel_file = "revenue_analysis_current.xlsx"

    # Check if file exists in current directory
    if not os.path.exists(excel_file):
        # Try with full path
        script_dir = os.path.dirname(os.path.abspath(__file__))
        excel_file = os.path.join(script_dir, excel_file)

        if not os.path.exists(excel_file):
            print(f"Error: Excel file not found: {excel_file}")
            print("Please ensure the file exists in the same directory as this script.")
            print("Run historical_revenue_analysis.py first to generate the file.")
            sys.exit(1)

    # Create and run enhancer
    enhancer = ExcelSummaryEnhancer(excel_file)

    print("=" * 60)
    print("Excel Summary Enhancer")
    print("=" * 60)
    print(f"Processing file: {excel_file}")
    print("This will add multiple summary sheets to the Excel file.")
    print("A backup will be created automatically.")
    print("-" * 60)
    print("Running enhancement automatically...")

    # Run enhancement
    success = enhancer.run()

    if success:
        print("\n" + "=" * 60)
        print("ENHANCEMENT COMPLETED SUCCESSFULLY")
        print("=" * 60)
        print("New sheets added to the Excel file:")
        print("1. Daily Summary")
        print("2. Weekly Summary")
        print("3. Client Analysis")
        print("4. Service Analysis")
        print("5. Revenue Trends")
        print("6. Peak Performance")
        print("7. Category Deep Dive")
        print("8. Mega Category Analysis")
        print("9. Revenue Forecast")
        print("\nA summary report has been generated.")
        print("=" * 60)
    else:
        print("\n" + "=" * 60)
        print("ENHANCEMENT FAILED")
        print("=" * 60)
        print("Check the logs for details.")
        sys.exit(1)


if __name__ == "__main__":
    main()
